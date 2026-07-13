from openpyxl import load_workbook

planilha = "classificacao_tributaria_organizada-aliquota.xlsx"

# ler a planilha
wb = load_workbook(planilha)    

# nome da aba
aba_cst = wb["Anexos"]

# start do dicionario
dados_planilha = {}
for linha in aba_cst.iter_rows(min_row=2, values_only=True):
    
    ncm = str(linha[3]).strip()
    
    
    if not ncm or ncm == "None" or len(ncm) == 9:
        continue
    
    
    dados_planilha[ncm] = {
        "cst": linha[0],
        "cClassTrib": linha[1],
        "descRed": linha[2],
        
        "redIBSMUN": linha[4],
        "redCBS": linha[5],
        "redIBSUF": linha[6],
        
        "difIBSMUN": linha[7],
        "difCBS": linha[8],
        "difIBSUF": linha[9],
        
        "aliqMonIBS": linha[10],
        "aliqMonCBS": linha[11],
        "fatorQtdBCMon": linha[12]
    }
    # print(dados_planilha[ncm], "\n")
              
print("ncms carregados:", len(dados_planilha))
# print(linha)
