from openpyxl import load_workbook

planilha = "classificacao_tributaria_organizada-cst.xlsx"

# ler a planilha
wb = load_workbook(planilha)    

# nome da aba
aba_cst = wb["Anexos"]

# start do dicionario
# start do dicionario
dados_planilha = {}
for linha in aba_cst.iter_rows(min_row=2, values_only=True):
    
    # Remove pontos, traços e espaços em branco que possam inflar o tamanho do NCM
    ncm = str(linha[4]).replace(".", "").replace("-", "").strip() if linha[3] is not None else ""
    
    # Se o NCM estiver vazio ou for a palavra literal "None", pula
    if not ncm or ncm == "None":
        continue
        
    dados_planilha[ncm] = {
        "cst": str(linha[0]).strip(),
        "cClassTrib": str(linha[1]).strip(),
        "descRed": linha[2],
    }
    
    
    print(dados_planilha[ncm], "\n")
              
print("ncms carregados:", len(dados_planilha))
# print(linha)
